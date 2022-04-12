from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Read write save
# wb = load_workbook('/media/muntasir/My Projects &  others/Python_Automation_projects/Excel11automation/Grades.xlsx')
# ws = wb.active 
# # print(ws['A1'].value)
# # ws['A2'].value = "Joy"
# # wb.save('/media/muntasir/My Projects &  others/Python_Automation_projects/Excel11automation/grades.xlsx') 
# print(wb.sheetnames)

# Creating a new xlsx file
# wb = Workbook()
# ws = wb.active
# ws.title = "Data"

# ws.append(['tim','is','Great','!'])
# wb.save('/media/muntasir/My Projects &  others/Python_Automation_projects/Excel11automation/new.xlsx') 

#Loop through data
# wb = load_workbook('/media/muntasir/My Projects &  others/Python_Automation_projects/Excel11automation/new.xlsx')
# ws = wb.active

# for row in range(1,11):
#     for col in range(1,5):
#         char = get_column_letter(col)
#         # print(ws[char +str(row)].value)
#         ws[char +str(row)] = char +str(row)
        
# wb.save("/media/muntasir/My Projects &  others/Python_Automation_projects/Excel11automation/new.xlsx")

data = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(data['Joe'].keys())
ws.append(headings)

for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)
    
for col in range(2, len(data['Joe'])+2):
    char = get_column_letter(col)
    ws[char+"7"] = f"=SUM({char+'2'}:{char+'6'})/{len(data)}"
    
for col in range(1, 6):
    ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")
    
wb.save("/media/muntasir/My Projects &  others/Python_Automation_projects/Excel11automation/Newgrades.xlsx")